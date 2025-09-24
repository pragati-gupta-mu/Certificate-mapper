
import streamlit as st
import openpyxl
from agent_api import process_rows_with_progress

def main():
    st.title("Certificate Mapper Agent UI")
    st.write("Upload your Excel file, select columns to send to the agent, and choose columns to update.")

    uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])
    if uploaded_file:
        wb = openpyxl.load_workbook(uploaded_file)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            st.error("No rows found in the uploaded file.")
            return
        headers = rows[0]
        st.write("Columns in your file:", headers)
        columns_to_send = st.multiselect(
            "Select columns to send to agent:", 
            headers, 
            default=["URL", "Certifier", "Certification Name"]
        )

        # Find default indices for selectboxes
        new_cert_index = 0
        remark_index = 0
        
        # Try to find columns that might be the new certificate or remark columns
        for i, header in enumerate(headers):
            if "NEW CERTIFICATE" in str(header).upper() or "CERTIFICATE NAME" in str(header).upper():
                new_cert_index = i
            elif "REMARK" in str(header).upper():
                remark_index = i

        new_cert_header = st.selectbox(
            "Select the new certificate name column:", headers, index=new_cert_index
        )

        remark_header = st.selectbox(
            "Select the remark column:", headers, index=remark_index
        )
        start_row = st.number_input("Start row (inclusive)", min_value=2, value=2, step=1)
        end_row = st.number_input("End row (inclusive)", min_value=start_row, value=start_row, step=1)
        row_numbers = ",".join(str(i) for i in range(start_row, end_row + 1))
        if st.button("Run Agent on Selected Rows"):
            try:
                row_nums = [int(x.strip()) for x in row_numbers.split(",") if x.strip().isdigit()]
                st.write(f"Rows to update: {row_nums}")
                
                # Prepare all row data for parallel processing
                rows_to_process = []
                row_indices = []
                
                for idx in row_nums:
                    if idx < 2 or idx > len(rows):
                        st.warning(f"Row {idx} is out of range.")
                        continue
                    row = rows[idx-1]
                    row_dict = {header: value for header, value in zip(headers, row) if header in columns_to_send}
                    rows_to_process.append(row_dict)
                    row_indices.append(idx)
                
                if rows_to_process:
                    st.write(f"Processing {len(rows_to_process)} rows in parallel...")
                    
                    # Add progress bar
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    def update_progress(completed, total):
                        progress = completed / total
                        progress_bar.progress(progress)
                        status_text.text(f"Processing: {completed}/{total} rows completed")
                    
                    # Process all rows in parallel with progress tracking
                    results = process_rows_with_progress(
                        rows_to_process, 
                        max_workers=3,  # Adjust based on your rate limits
                        progress_callback=update_progress
                    )
                    
                    # Update Excel file with results
                    for idx, result in zip(row_indices, results):
                        st.write(f"Row {idx} result: {result}")
                        ws.cell(row=idx, column=headers.index(new_cert_header) + 1, value=result.get("newCertificateName"))
                        ws.cell(row=idx, column=headers.index(remark_header) + 1, value=result.get("remark"))
                    
                    progress_bar.progress(1.0)
                    status_text.text("✅ All rows processed successfully!")
                    
                # Save updated file
                output_path = "updated_" + uploaded_file.name
                wb.save(output_path)
                st.success(f"Updated file saved as {output_path}")
                with open(output_path, "rb") as f:
                    st.download_button("Download updated file", f, file_name=output_path)
            except Exception as e:
                st.error(f"Error: {e}")

if __name__ == "__main__":
    main()
